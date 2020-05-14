import kivy
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.core.audio import SoundLoader
from kivy.base import runTouchApp
from kivy.uix.button import Button
from kivy.uix.popup import Popup
import time
from kivy.config import Config
from kivy.core.window import Window
import pandas as pd
import matplotlib.pyplot as plt
from pandas import ExcelWriter
from pandas import ExcelFile
import pandas as pd 
import pprint
import openpyxl
from pathlib import Path
import easygui
import heapq
import random
import xlrd
from easygui import choicebox

Config.set('graphics', 'resizable', True)
  
Window.clearcolor = (1, 1, 1, 1) 
x=random.choice(['red', 'black', 'green','orange','blue','pink','purple','brown','grey']) 


class MyGrid(GridLayout):
        
    

    def __init__(self, **kwargs):
        super(MyGrid, self).__init__(**kwargs)
        self.cols=1
        self.btn=Button(background_normal = '/Users/satyamkumar/Desktop/3.png', background_down ='/Users/satyamkumar/Desktop/3.png')
        self.add_widget(self.btn)

        self.inside=GridLayout()
        self.inside.cols=3
        
        self.inside.add_widget(Label(text="[b] HIGH RISK TRAVEL \ndestination countries:\n (please wait after clicking)[/b]",markup=True ,color=(0,0,0,1)))
        self.btn1=Button(text ="click here", color =(0, 0, 0, 1), background_normal = '/Users/satyamkumar/Desktop/1.png', background_down ='/Users/satyamkumar/Desktop/2.png', size_hint = (.1, .1), pos_hint = {"x":0.35, "y":0.3} )
        self.btn1.bind(on_press = self.pressed1)
        self.btn4=Button(text="Graph 1",color =(0, 0, 0, 1),background_normal = '/Users/satyamkumar/Desktop/8.1.jpg')
        self.btn4.bind(on_press=self.pressed4)
        self.inside.add_widget(self.btn1)
        self.inside.add_widget(self.btn4)
        

        self.inside.add_widget(Label(text="[b]   TOP 5 countries \nas HIGH RISK for Study  :\n  (graph for 11)  [/b] ",markup=True,color=(0,0,0,1)))
        self.btn2=Button(text="click here",color =(0, 0, 0, 1),background_normal = '/Users/satyamkumar/Desktop/4.jpg', background_down ='/Users/satyamkumar/Desktop/5.jpg')
        self.btn2.bind(on_press = self.pressed2)
        self.btn5=Button(text="Graph 2",color =(0, 0, 0, 1),background_normal = '/Users/satyamkumar/Desktop/8.1.jpg')
        self.btn5.bind(on_press = self.pressed5)
        self.inside.add_widget(self.btn2)
        self.inside.add_widget(self.btn5)
        

        self.inside.add_widget(Label(text="[b]   Average days for \nDEATHS and RECOVERIES\n(for country wise graph\n press 2nd button) :  [/b]",markup=True ,color=(0,0,0,1)))
        self.btn3=Button(text="click here",color =(0, 0, 0, 1),background_normal = '/Users/satyamkumar/Desktop/6.jpg', background_down ='/Users/satyamkumar/Desktop/7.jpg')
        self.btn3.bind(on_press=self.pressed3)
        self.btn6=Button(text="Graph 3",color =(0, 0, 0, 1),background_normal = '/Users/satyamkumar/Desktop/8.1.jpg')
        self.btn6.bind(on_press = self.pressed6)
        self.inside.add_widget(self.btn3)
        self.inside.add_widget(self.btn6)

        
        self.add_widget(self.inside)
        
    def pressed1(self,instance):
        dfconfirmed = pd.read_excel(r"/Users/satyamkumar/Desktop/data8.1.xlsx", sheet_name=0)
        dfdeath = pd.read_excel(r"/Users/satyamkumar/Desktop/data8.2.xlsx",sheet_name=0 )
        dfconfirmed= dfconfirmed.loc[:,'1/22/20':]
        dfdeath= dfdeath.loc[:,'1/22/20':]
        index = list(dfdeath.index)
        RiskPlaces = list()
        for i in index:
            for j, x in zip(dfconfirmed, dfdeath):
                if (dfdeath.loc[i][x] == 0) | (dfconfirmed.loc[i][j] == 0):
                    case_fatality = 0
                elif dfconfirmed.loc[i][j] != 0:
                    case_fatality = (dfdeath.loc[i][x] / dfconfirmed.loc[i][j]) * 100
                else:
                    case_fatality = 0
                y = (4 / 100)*dfconfirmed.loc[i][j]

                if ((case_fatality != 0) & (case_fatality >= y)):
                    if i in RiskPlaces:
                        break
                    else:
                        RiskPlaces.append(i)
        list1=[]
        dfconfirmed = pd.read_excel(r"/Users/satyamkumar/Desktop/data8.1.xlsx", sheet_name=0)
        for i in range(0,81):
            if(i==0):
                print("Afghanistan")
            else:
                x = dfconfirmed['Country'][RiskPlaces[i-1]]
                list1.append(x)
        layout=GridLayout(cols=5 ,padding=10)
        for i in range (0,80):       
            popupLabel=Label(text=list1[i])
            layout.add_widget(popupLabel)
        popup=Popup(title="Top countries at high risk for next two years:",content=layout,size_hint=(None,None),size=(600,300),background='/Users/satyamkumar/Desktop/9.jpg')
        popup.open()               
        
    
    def pressed2(self,instance):
        xlsx_file = Path("/Users/satyamkumar/Desktop/data6.xlsx")

        wb_obj = openpyxl.load_workbook(xlsx_file)

        sheet = wb_obj.active

        conformed_cases = []
        i = 0
        cases = 0
        for row in sheet.iter_rows(2,sheet.max_row):
            if(i % 2 == 0):
                cases = row[3].value
                i += 1
            else:
                conformed_cases.append((row[3].value)+cases)
                i += 1
        list1=[]
        list = heapq.nlargest(5, range(len(conformed_cases)), key=conformed_cases.__getitem__)
        new_list = [x+1 for x in list]
        for i in new_list:
            for row in sheet.iter_rows(2,sheet.max_row):
                 if(row[0].value == i):
                     list1.append(row[1].value)
        
        layout=GridLayout(cols=1 ,padding=10)
        for i in range(0,5):
            popupLabel=Label(text=list1[i])
            layout.add_widget(popupLabel)
        popup=Popup(title="Top 5 dangerous states are:",content=layout,size_hint=(None,None),size=(600,300),background='/Users/satyamkumar/Desktop/9.jpg')
        popup.open()


    def pressed3(self,instance):
        xlsx_file = Path("/Users/satyamkumar/Desktop/data8.1.xlsx")
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active


        xlsx_file1 = Path("/Users/satyamkumar/Desktop/data8.2.xlsx")
        wb_obj1 = openpyxl.load_workbook(xlsx_file1)
        sheet1 = wb_obj1.active


        xlsx_file2 = Path("/Users/satyamkumar/Desktop/data8.3.xlsx")
        wb_obj2 = openpyxl.load_workbook(xlsx_file2)
        sheet2 = wb_obj2.active

        max_col = sheet.max_column
        max_rows = sheet.max_row

        list =[]
        list1 = []
        list2 = []
        list3 = []
        list4 = []
        list5 = []


        for i in range(4, max_col + 1): 
            temp = []
            for j in range(2, max_rows + 1):
                cell_obj = sheet.cell(row=j, column=i)
                temp.append(cell_obj.value)
            list.append(sum(temp))

        # for deaths
        for i in range(4, max_col + 1): 
            temp = []
            for j in range(2, max_rows + 1):
                cell_obj = sheet1.cell(row=j, column=i)
                temp.append(cell_obj.value)
            list2.append(sum(temp))

        # for recovery cases
        for i in range(4, max_col + 1): 
            temp = []
            for j in range(2, max_rows + 1):
                cell_obj = sheet2.cell(row=j, column=i)
                temp.append(cell_obj.value)
            list4.append(sum(temp))

        s = 0
        s1 = 0
        s2 = 0
        # for conformed cases
        for i in range(1,len(list)+1):
            if(i%7 == 0):
                s1 += list[i-1]
                list1.append(s1)
                s1 = 0
            else:
                s1 += list[i-1]

        # for death cases
        for i in range(1,len(list2)+1):
            if(i%7 == 0):
                s2 += list2[i-1]
                list3.append(s2)
                s2 = 0
            else:
                s2 += list2[i-1]

        # for recovery cases
        for i in range(1,len(list4)+1):
            if(i%7 == 0):
                s += list4[i-1]
                list5.append(s)
                s = 0
            else:
                s += list4[i-1]
        lista=[]
        listb=[]
        listc=[]
        
        j = 0
        for i in list1:
            j += 1
            lista.append("Cnf week "+str(j)+" : "+str(i//7))

        j = 0
        for i in list3:
            j += 1
            listb.append("Death week "+str(j)+" : "+str(i//7))

        j = 0
        for i in list5:
            j += 1
            listc.append("Rec. week "+str(j)+" : "+str(i//7))
        layout=GridLayout(cols=3,padding=20)
        for i in range(0,13):
            popupLabel=Label(text=lista[i])
            layout.add_widget(popupLabel)
    
        for i in range(0,13):
            popupLabel=Label(text=listb[i])
            layout.add_widget(popupLabel)
    
        for i in range(0,13):
            popupLabel=Label(text=listc[i])
            layout.add_widget(popupLabel)
        popup=Popup(title="Average cases per week :",content=layout,size_hint=(None,None),size=(600,300),background='/Users/satyamkumar/Desktop/9.jpg')
        popup.open()    



    def pressed4(self,instance):
        dfconfirmed = pd.read_excel(r"/Users/satyamkumar/Desktop/data8.1.xlsx", sheet_name=0)
        dfdeath = pd.read_excel(r"/Users/satyamkumar/Desktop/data8.2.xlsx",sheet_name=0 )
        dfconfirmed= dfconfirmed.loc[:,'1/22/20':]
        dfdeath= dfdeath.loc[:,'1/22/20':]
        index = list(dfdeath.index)
        RiskPlaces = list()
        for i in index:
            for j, x in zip(dfconfirmed, dfdeath):
                if (dfdeath.loc[i][x] == 0) | (dfconfirmed.loc[i][j] == 0):
                    case_fatality = 0
                elif dfconfirmed.loc[i][j] != 0:
                    case_fatality = (dfdeath.loc[i][x] / dfconfirmed.loc[i][j]) * 100
                else:
                    case_fatality = 0
                y = (4 / 100)*dfconfirmed.loc[i][j]

                if ((case_fatality != 0) & (case_fatality >= y)):
                    if i in RiskPlaces:
                        break
                    else:
                        RiskPlaces.append(i)
        
        list1=[]
        list2=[]
        dfconfirmed = pd.read_excel(r"/Users/satyamkumar/Desktop/data8.1.xlsx", sheet_name=0)
        for i in range(0,81):
            if(i==0):
                print("Afghanistan")
                list2.append(1092)
            else:
                x = dfconfirmed['Country'][RiskPlaces[i-1]]
                y = dfconfirmed['4/21/20'][RiskPlaces[i-1]]
                list1.append(x)
                list2.append(y)
        plt.plot(RiskPlaces,list2)       
        plt.xlabel('sno. of countries')
        plt.ylabel('cases on most recent date(21/04/20')
        plt.show()




    def pressed5(self,instance):
        
        a ="/Users/satyamkumar/Desktop/data6.1.xlsx"
        df=pd.read_excel(a)
        val=df[['Cities','Confirmed Cases']]
        val.plot.bar(x="Cities",y="Confirmed Cases",figsize=[9,9],color=x)
        plt.show() 

    def pressed6(self,instance):
        df = pd.read_excel(r"/Users/satyamkumar/Desktop/data7.1.xlsx",sheet_name=0)
        loc = ("/Users/satyamkumar/Desktop/data7.1.xlsx") 
        wb = xlrd.open_workbook(loc) 
        sheet = wb.sheet_by_index(0)
        myvar = easygui.enterbox("Enter the country name you want to view\n(first letter capital) ")
        for i in range(0,186):
            if(df['Country/Region'][i]==myvar):
       
                x=random.choice(['red', 'black', 'green','orange','blue','pink','purple','brown','grey']) 
                plt.figure(figsize=[90,180])
                plt.plot(sheet.row_values(i+1),color=x)
                plt.xlabel("number of days")
                plt.ylabel("counts of cases")
                plt.ylim(0,70)
                plt.show()
                break
    
            else:
                pass                 
             




class TrackitApp(App):
    def build(self):
        
       return MyGrid()

if __name__=="__main__":
    TrackitApp().run()
