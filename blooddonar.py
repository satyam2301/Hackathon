#python code to make or udate data of user for blood donation
#hackathon 1
#Team Member : PRIYAM (19BCS089) and SATYAM KUMAR (19BEC040)
# exact address of the file has to be placed in line number 10,15,116 and 117

import openpyxl
from pathlib import Path

# Donar file
xlsx_file = Path("/Users/satyamkumar/Desktop/donar.xlsx")
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active

# Reciepient file
xlsx_file1 = Path("/Users/satyamkumar/Desktop/reciepient.xlsx")
wb_obj1 = openpyxl.load_workbook(xlsx_file1)
sheet1 = wb_obj1.active

while True:
    donar=[]
    recipient=[] 
    a = sheet.max_row
    a += 1
    b = 1
    c = sheet1.max_row
    c += 1
    d = 1
    print('\n1. Enter Donor data\n2. Enter Recipient data\n3. Get Donor data \
    \n4. Get Recipient data\n5. Exit From The system')
    ch = int(input('\nEnter your choice : '))

    if ch==1:
        print('\nEnter the following details of the Donar: ')
        id = input("\nDonar ID: ")
        donar.append(id)
        name = input("\nName: ")
        donar.append(name)
        bg = input("\nBloodgroup: ")
        donar.append(bg)
        age = input("\nAge: ")
        donar.append(age)
        gender = input("\nGender: ")
        donar.append(gender)
        pno = input("\nPhone Number: ")
        donar.append(pno)
        eid = input("\nEmail ID: ")
        donar.append(eid)
        address = input("\nAddress: ")
        donar.append(address)
        for data in donar:
            temp = sheet.cell(row = a,column = b)
            temp.value = data
            b += 1
        print("\n\nThank You!!\nNew Donor Registered...")
        a += 1
        donar = []

    elif ch==2:
        print("\nEnter following details: ")
        id = input("\nReciepient ID: ")
        reciepient.append(id)
        name = input("\nName: ")
        reciepient.append(name)
        bg = input("\nBloodgroup: ")
        reciepient.append(bg)
        age = input("\nAge: ")
        reciepient.append(age)
        gender = input("\nGender: ")
        reciepient.append(gender)
        pno = input("\nPhone Number: ")
        reciepient.append(pno)
        eid = input("\nEmail ID: ")
        reciepient.append(eid)
        address = input("\nAddress: ")
        reciepient.append(address)  
        for data in reciepient:
            temp = sheet1.cell(row = c,column = d)
            temp.value = data
            d += 1
        print("\n\nThank You!!\nNew Reciepient Registered...")
        c += 1
        reciepient = []

    elif ch==3:
        Did = input("Enter Doner ID: ")
        x = 0
        for i in range(1,sheet.max_row+1):
            cell_obj = sheet.cell(row = i,column = 1)
            if cell_obj.value == Did:
                for j in range(1,sheet.max_column+1):
                    cell_obj1 = sheet.cell(row =i,column =j)
                    print(cell_obj1.value,end="\t")
                x = 1
        if x == 0:
            print("Donar ID not found!!!")
        print("\n")

    elif ch==4:
        Rid = input("Enter Reciepient ID: ")
        x = 0
        for i in range(1,sheet1.max_row+1):
            cell_obj = sheet1.cell(row = i,column = 1)
            if cell_obj.value == Rid:
                for j in range(1,sheet1.max_column+1):
                    cell_obj1 = sheet1.cell(row =i,column =j)
                    print(cell_obj1.value,end="\t")
                x = 1
        if x == 0:
            print("Reciepient ID not found!!!")
        print("\n")
    elif ch==5:
        exit()
    else:
        print("\n\nSorry Wrong Option!!\n We Regret For This Inconvenience Caused!!\n----Try Again!!----")

    wb_obj.save("/Users/satyamkumar/Desktop/donar.xlsx")
    wb_obj1.save("/Users/satyamkumar/Desktop/reciepient.xlsx")